#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch Hybrid Ozon Parser - БЫСТРАЯ ВЕРСИЯ без прокси для теста

Для быстрого тестирования работает БЕЗ прокси.
После успешного теста можно вернуть прокси.
"""
import json
import time
import os
import re
from datetime import datetime
from curl_cffi.requests import Session as CffiSession
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIGURATION ---
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
API_ENDPOINT = "https://www.ozon.ru/api/composer-api.bx/page/json/v2"
USE_PROXY = False  # Отключаем прокси для быстрого теста

def warmup_session_once():
    """Прогрев сессии без прокси для быстрого теста."""
    options = uc.ChromeOptions()
    options.binary_location = CHROME_PATH
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=ru-RU")
    # БЕЗ ПРОКСИ для быстрого теста
    
    driver = None
    try:
        print("[WARMUP] Запуск браузера (БЕЗ ПРОКСИ - быстрый тест)...")
        driver = uc.Chrome(options=options, browser_executable_path=CHROME_PATH)
        
        print("[WARMUP] Посещение главной страницы Ozon...")
        driver.get("https://www.ozon.ru")
        time.sleep(3)
        
        print("[WARMUP] Прогрев на странице товара...")
        driver.get("https://www.ozon.ru/product/1067025156/")
        time.sleep(5)
        
        selenium_cookies = driver.get_cookies()
        cookies_dict = {cookie['name']: cookie['value'] for cookie in selenium_cookies}
        user_agent = driver.execute_script("return navigator.userAgent;")
        
        print(f"[WARMUP] ✅ Сессия прогрета. Извлечено {len(cookies_dict)} куки.")
        
        return cookies_dict, user_agent
        
    except Exception as e:
        print(f"[ERROR] Ошибка прогрева: {e}")
        return None, None
    finally:
        if driver:
            driver.quit()
            print("[WARMUP] Браузер закрыт.")

def clean_price(price_str):
    """Очистка строки цены от лишних символов (валюта, пробелы)."""
    if not price_str:
        return None
    # Оставляем только цифры
    cleaned = re.sub(r'[^\d]', '', str(price_str))
    return int(cleaned) if cleaned else None

def fetch_product_batch(product_links, cookies_dict, user_agent):
    """Обработка пакета товаров через API с извлечением цен и статуса наличия."""
    session = CffiSession(impersonate="chrome124")
    
    headers = {
        "authority": "www.ozon.ru",
        "accept": "application/json",
        "accept-language": "ru-RU,ru;q=0.9",
        "user-agent": user_agent,
        "x-o3-app-name": "entrypoint-api",
        "x-o3-app-version": "master",
        "sec-ch-ua": '"Not_A Brand";v="124", "Chromium";v="124", "Google Chrome";v="124"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }
    
    results = []
    
    for idx, product_link in enumerate(product_links, 1):
        try:
            print(f"[{idx}/{len(product_links)}] Обработка: {product_link}")
            
            payload = {"url": product_link}
            headers["referer"] = f"https://www.ozon.ru{product_link}"
            
            response = session.get(
                API_ENDPOINT,
                params=payload,
                headers=headers,
                cookies=cookies_dict,
                timeout=15
            )
            
            if response.status_code != 200:
                print(f"  ❌ Ошибка: HTTP {response.status_code}")
                results.append({
                    "link": product_link,
                    "status": f"ERROR_{response.status_code}",
                    "title": None,
                    "price_card": None,
                    "price_standard": None,
                    "price_old": None,
                    "stock_status": "Error"
                })
                continue
            
            data = response.json()
            seo_data = data.get("seo") or data.get("SEO")
            widget_states = data.get("widgetStates", {})
            
            # Если widgetStates пуст на верхнем уровне, ищем глубже (новая структура Ozon)
            if not widget_states:
                try:
                    # Путь для Nuxt/Composer структуры
                    vi = data.get("verticalInfo", {})
                    composer = vi.get("composer", {}) or vi.get("pdp", {})
                    widget_states = composer.get("widgetStates", {})
                except:
                    pass

            title = "Unknown"
            if seo_data:
                title = seo_data.get("title") or "Unknown"
            
            # Извлечение цен из виджетов
            price_card = None
            price_standard = None
            price_old = None
            stock_status = "In Stock"
            
            # Ищем виджет цены или OOS
            price_widget_key = next((k for k in widget_states.keys() if "webPrice" in k), None)
            oos_widget_key = next((k for k in widget_states.keys() if "webOutOfStock" in k), None)
            
            if price_widget_key:
                try:
                    price_state = json.loads(widget_states[price_widget_key])
                    price_card = clean_price(price_state.get("cardPrice"))
                    price_standard = clean_price(price_state.get("price"))
                    price_old = clean_price(price_state.get("originalPrice"))
                    
                    # Проверка на статус OOS внутри webPrice (иногда там)
                    if "закончился" in str(price_state).lower():
                        stock_status = "Out of Stock"
                except Exception as e:
                    print(f"  ⚠️ Ошибка парсинга виджета цен: {e}")

            if oos_widget_key:
                 stock_status = "Out of Stock"
                 try:
                     oos_state = json.loads(widget_states[oos_widget_key])
                     # Иногда старая цена сохраняется в OOS виджете
                     if not price_standard:
                         price_standard = clean_price(oos_state.get("price"))
                 except:
                     pass
            
            # Дополнительная проверка по тексту если виджеты не дали явного ответа
            if stock_status == "In Stock" and price_standard is None:
                if "закончился" in str(data).lower():
                    stock_status = "Out of Stock"
            
            # Фолбэк на SEO данные если виджет не найден (только для основной цены)
            if price_standard is None and seo_data:
                price_standard = clean_price(seo_data.get("price"))
            
            print(f"  ✅ {title[:40]}...")
            print(f"     [Статус] {stock_status}")
            print(f"     [Цены] Карта: {price_card} | Стандарт: {price_standard} | Старая: {price_old}")
            
            results.append({
                "link": product_link,
                "status": "OK",
                "title": title,
                "price_card": price_card,
                "price_standard": price_standard,
                "price_old": price_old,
                "stock_status": stock_status,
                "currency": "RUB"
            })
            
            time.sleep(0.5)
            
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            results.append({
                "link": product_link,
                "status": "EXCEPTION",
                "title": None,
                "price_card": None,
                "price_standard": None,
                "price_old": None,
                "stock_status": "Exception",
                "error": str(e)
            })
    
    return results

def save_results(results, filename="hybrid_results"):
    """Сохранение результатов в JSON и Excel."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    json_file = f"{filename}_{timestamp}.json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\n💾 Сохранено в JSON: {json_file}")
    
    excel_file = f"{filename}_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    headers = ["№", "Ссылка", "Статус Парсинга", "Наличие", "Название", "Цена (Карта)", "Цена (Стандарт)", "Цена (Старая)", "Валюта"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, result in enumerate(results, 1):
        ws.append([
            idx,
            result.get("link", ""),
            result.get("status", ""),
            result.get("stock_status", ""),
            result.get("title", ""),
            result.get("price_card", ""),
            result.get("price_standard", ""),
            result.get("price_old", ""),
            result.get("currency", "")
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_file)
    print(f"💾 Сохранено в Excel: {excel_file}")
    
    return json_file, excel_file

def process_batch(product_links):
    """Главная функция пакетной обработки."""
    print("="*70)
    print(f" HYBRID BATCH PARSER - {len(product_links)} товаров (БЕЗ ПРОКСИ)")
    print("="*70)
    
    start_time = time.time()
    
    cookies, user_agent = warmup_session_once()
    
    if not cookies or not user_agent:
        print("❌ Не удалось прогреть сессию!")
        return
    
    results = fetch_product_batch(product_links, cookies, user_agent)
    json_file, excel_file = save_results(results)
    
    elapsed = time.time() - start_time
    success_count = sum(1 for r in results if r.get("status") == "OK")
    error_count = len(results) - success_count
    
    print("\n" + "="*70)
    print(" ИТОГИ ")
    print("="*70)
    print(f"✅ Успешно:  {success_count}/{len(results)}")
    print(f"❌ Ошибки:   {error_count}/{len(results)}")
    print(f"⏱️  Время:    {elapsed:.1f}s ({elapsed/len(results):.1f}s на товар)")
    print(f"📊 Скорость:  {len(results)/(elapsed/60):.1f} товаров/мин")
    print("="*70)

if __name__ == "__main__":
    # Приоритет 1: Все товары из БД
    if os.path.exists("all_products_from_db.txt"):
        with open("all_products_from_db.txt", "r", encoding="utf-8") as f:
            product_links = [line.strip() for line in f if line.strip()]
        print(f"📂 Загружено {len(product_links)} товаров из БД\n")
    # Приоритет 2: Тестовый список
    elif os.path.exists("test_products_10.txt"):
        with open("test_products_10.txt", "r", encoding="utf-8") as f:
            product_links = [line.strip() for line in f if line.strip()]
        print(f"📂 Загружено {len(product_links)} тестовых товаров\n")
    # Приоритет 3: Хардкод
    else:
        product_links = [
            "/product/1067025156/",
            "/product/1564586312/",
            "/product/1401683802/",
        ]
        print(f"⚠️  Используется тестовый список: {len(product_links)} товаров\n")
    
    if len(product_links) > 0:
        process_batch(product_links)
    else:
        print("❌ Список товаров пуст!")
