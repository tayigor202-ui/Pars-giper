import os
import psycopg2
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
DB_URL = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASS')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"

print("="*70)
print("ТЕСТ НОВОГО ФОРМАТА ОТЧЁТА")
print("="*70)

conn=psycopg2.connect(DB_URL)
query="""SELECT sku,name,competitor_name,price_card,price_nocard,price_old,status,sp_code FROM public.prices ORDER BY name, competitor_name"""
df=pd.read_sql(query,conn)
conn.close()

# Apply mapping
store_mapping={
    'Ссылка на наш магазин':'Наш магазин',
    'Магазин DeLonghi Group':'DeLonghi Group',
    'DeLonghi Group':'DeLonghi Group',
    'Delonghi Official Store':'DeLonghi Official',
    'Delonghi official store':'DeLonghi Official',
    'DeLonghi Official Store':'DeLonghi Official'
}
df['competitor_name']=df['competitor_name'].map(lambda x:store_mapping.get(x,x))

# Status transformation
def fill_status(row):
    status = row.get('status')
    p_card = row.get('price_card')
    
    if status == 'OUT_OF_STOCK':
        return pd.Series(['Товар закончился', 'Товар закончился', 'Товар закончился'], index=['price_card', 'price_nocard', 'price_old'])
    
    if pd.isna(p_card) or str(p_card).lower().strip() in ['','none','nan']:
        text = 'Нет данных'
        return pd.Series([text, text, text], index=['price_card', 'price_nocard', 'price_old'])
        
    return pd.Series([row['price_card'], row['price_nocard'], row['price_old']], index=['price_card', 'price_nocard', 'price_old'])

df[['price_card', 'price_nocard', 'price_old']] = df.apply(fill_status, axis=1)

# New pivot: ONLY sp_code as index
pivot_df = df.pivot_table(
    index='sp_code',
    columns='competitor_name', 
    values=['name', 'sku', 'price_card', 'price_nocard', 'price_old'],
    aggfunc='first',
    dropna=False
)

pivot_df.columns = pivot_df.columns.swaplevel(0, 1)

rename_map = {
    'name': 'Название',
    'sku': 'SKU',
    'price_card': 'Цена с картой',
    'price_nocard': 'Цена без карты',
    'price_old': 'Старая цена'
}
pivot_df = pivot_df.rename(columns=rename_map, level=1)

for col in pivot_df.columns:
    if col[1] in ['Цена с картой', 'Цена без карты', 'Старая цена']:
        pivot_df[col] = pivot_df[col].fillna('Нет данных')

sellers = sorted(pivot_df.columns.get_level_values(0).unique())
desired_order = ['Название', 'SKU', 'Цена с картой', 'Цена без карты', 'Старая цена']

new_columns = []
for seller in sellers:
    for attr in desired_order:
        if (seller, attr) in pivot_df.columns:
            new_columns.append((seller, attr))

pivot_df = pivot_df.reindex(columns=new_columns)

print(f"\n✅ РЕЗУЛЬТАТ:")
print(f"   Строк (SP-CODE): {len(pivot_df)}")
print(f"   Колонок: {len(pivot_df.columns)}")
print(f"   Магазинов: {len(sellers)}")

print(f"\n📊 ПЕРВЫЕ 3 СТРОКИ:")
print(pivot_df.head(3))

# Save test report
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"test_report_fixed_{timestamp}.xlsx"

with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    pivot_df.to_excel(writer, sheet_name='Цены')
    worksheet = writer.sheets['Цены']
    
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    
    for i, column in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    worksheet.freeze_panes = 'B3'

print(f"\n✅ Тестовый отчёт создан: {filename}")
print("="*70)
