#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys, time, random, json, re, threading, requests, shutil, string, subprocess, psutil
from queue import Queue
from datetime import datetime, timedelta
from pathlib import Path

import psycopg2
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

import undetected_chromedriver as uc
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dotenv import load_dotenv
import pandas as pd
from curl_cffi.requests import Session as CffiSession

API_ENDPOINT = "https://www.ozon.ru/api/composer-api.bx/page/json/v2"

# Optional: scheduler module
try:
    import scheduler
except ImportError:
    scheduler = None

# Optional: check_violations module
try:
    import check_violations
except ImportError:
    check_violations = None

load_dotenv()

ip_timezone_cache={}
ip_cache_lock=threading.Lock()

def get_ip_geolocation(ip):
    try:
        response=requests.get(f'http://ip-api.com/json/{ip}?fields=status,country,city,timezone',timeout=5)
        if response.status_code==200:
            data=response.json()
            if data.get('status')=='success':
                return {'country':data.get('country','Russia'),'city':data.get('city','Moscow'),'timezone':data.get('timezone','Europe/Moscow')}
    except:
        pass
    return {'country':'Russia','city':'Moscow','timezone':'Europe/Moscow'}

def get_timezone_offset(timezone_name):
    timezone_offsets={'Europe/Moscow':-180,'Europe/Kaliningrad':-120,'Europe/Samara':-240,'Asia/Yekaterinburg':-300,'Asia/Omsk':-360,'Asia/Krasnoyarsk':-420,'Asia/Irkutsk':-480,'Asia/Yakutsk':-540,'Asia/Vladivostok':-600,'Asia/Magadan':-660,'Asia/Kamchatka':-720}
    return timezone_offsets.get(timezone_name,-180)

def get_timezone_for_ip(ip):
    with ip_cache_lock:
        if ip in ip_timezone_cache:
            return ip_timezone_cache[ip]
    geo=get_ip_geolocation(ip)
    timezone_name=geo.get('timezone','Europe/Moscow')
    offset=get_timezone_offset(timezone_name)
    with ip_cache_lock:
        ip_timezone_cache[ip]={'offset':offset,'name':timezone_name,'city':geo.get('city','Moscow')}
    return ip_timezone_cache[ip]


DB_URL=f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASS')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
TG_BOT_TOKEN=os.getenv('TG_BOT_TOKEN')
TG_CHAT_ID=os.getenv('TG_CHAT_ID')
CHROME_PATH=r"C:\Program Files\Google\Chrome\Application\chrome.exe"
DEBUG_PORT_START=9222
NUM_WORKERS=1
BATCH_SIZE=180
USE_HEADLESS=False
MAX_PRODUCTS_PER_BATCH=180
RESUME_FROM_LAST_N=0
DELAY_BETWEEN_PRODUCTS=(3.0,7.0)
BATCH_PAUSE_INTERVAL=20
BATCH_PAUSE_DURATION=(10.0,30.0)
MAX_RETRIES_PER_PRODUCT=3


product_queue=Queue()
results=[]
results_lock=threading.Lock()
db_save_counter=0
db_save_lock=threading.Lock()
processed_count=0
processed_lock=threading.Lock()
stop_flag=False
retry_queue=Queue()
retry_counts={}
retry_lock=threading.Lock()
batch_complete=False
batch_lock=threading.Lock()
last_processed_skus=[]
antibot_detected=False
antibot_lock=threading.Lock()

global_cookies = None
global_ua = None

def warmup_session():
    """Прогрев сессии напрямую (без прокси) для получения актуальных куки и User-Agent."""
    options = uc.ChromeOptions()
    options.binary_location = CHROME_PATH
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=ru-RU")
    
    if USE_HEADLESS:
        options.add_argument("--headless=new")
    
    # Добавляем случайный UA
    ua = generate_random_user_agent()
    options.add_argument(f"user-agent={ua}")
    
    driver = None
    try:
        print("[WARMUP] Запуск браузера (ПРЯМОЕ СОЕДИНЕНИЕ - быстрый тест)...")
        driver = uc.Chrome(options=options, browser_executable_path=CHROME_PATH)
        
        print("[WARMUP] Посещение Ozon...")
        driver.get("https://www.ozon.ru")
        time.sleep(3)
        
        # Прогрев на странице любого товара (как в тесте)
        print("[WARMUP] Прогрев на странице товара...")
        driver.get("https://www.ozon.ru/product/1067025156/")
        time.sleep(5)
            
        selenium_cookies = driver.get_cookies()
        cookies_dict = {cookie['name']: cookie['value'] for cookie in selenium_cookies}
        user_agent = driver.execute_script("return navigator.userAgent;")
        
        print(f"[WARMUP] ✅ Сессия прогрета. Извлечено {len(cookies_dict)} куки.")
        return cookies_dict, user_agent
    except Exception as e:
        print(f"[WARMUP] ERROR: {e}")
        return None, None
    finally:
        if driver:
            try: driver.quit()
            except: pass

def clean_price(price_str):
    if not price_str:
        return None
    cleaned = re.sub(r'[^\d]', '', str(price_str))
    return int(cleaned) if cleaned else None

def generate_random_user_agent_full():
    chrome_versions = [
        ('144.0.7559.59', '144'), ('131.0.6778.85', '131'), 
        ('130.0.6723.116', '130'), ('129.0.6668.100', '129'),
        ('128.0.6613.138', '128'), ('127.0.6533.119', '127')
    ]
    edge_versions = [
        ('131.0.2903.70', '131'), ('130.0.2849.80', '130'),
        ('129.0.2792.89', '129')
    ]
    
    is_edge = random.random() < 0.2
    if is_edge:
        ver, major = random.choice(edge_versions)
        ua = f'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{ver} Safari/537.36 Edg/{ver}'
        brand = '"Microsoft Edge";v="%s", "Chromium";v="%s", "Not?A_Brand";v="99"' % (major, major)
    else:
        ver, major = random.choice(chrome_versions)
        ua = f'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{ver} Safari/537.36'
        brand = '"Google Chrome";v="%s", "Chromium";v="%s", "Not?A_Brand";v="99"' % (major, major)
        
    return {
        'ua': ua, 
        'major': major, 
        'brand': brand, 
        'full_ver': ver, 
        'platform': 'Windows',
        'platform_version': '13.0.0', # Win 11 modern
        'architecture': 'x86',
        'model': '',
        'bitness': '64'
    }

def generate_random_user_agent():
    return generate_random_user_agent_full()['ua']

def start_browser_uc(port, unique_id, ua_info, proxy_host, proxy_port, proxy_user, proxy_pass, worker_id):
    # USE UNIQUE TEMPORARY PROFILE
    profile=f"C:\\Temp\\chrome_profiles\\ozon\\tmp_{unique_id}"
    Path(profile).mkdir(parents=True,exist_ok=True)
    
    user_agent = ua_info['ua']
    brand = ua_info['brand']
    
    options = uc.ChromeOptions()
    # Basics
    options.add_argument(f"--user-data-dir={profile}")
    options.add_argument(f"--user-agent={user_agent}")
    options.add_argument(f"--proxy-server=http://{proxy_host}:{proxy_port}")
    
    # Anti-Fingerprinting & Privacy
    # Remove --disable-features and --disable-web-security as they are detectable
    options.add_argument("--lang=ru-RU")
    
    # WebRTC Protection
    options.add_argument("--disable-webrtc")
    options.add_argument("--enforce-webrtc-ip-permission-check")
    
    # Performance & Stealth
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")
    
    if USE_HEADLESS:
        # Use the newer, more stealthy headless mode if requested
        options.add_argument("--headless=new")
    
    # Resolution & Hiding
    desktop_resolutions=['1920,1080','1366,768','1536,864','1440,900','1600,900','1280,720']
    res = random.choice(desktop_resolutions)
    options.add_argument(f"--window-size={res}")
    # Diagnostic: Move window to (0,0) to see if off-screen was the trigger
    options.add_argument("--window-position=0,0")

    max_start_retries = 3
    driver = None
    for attempt in range(max_start_retries):
        try:
            print(f"[W{worker_id}] 🚀 Attempt {attempt+1} to start browser on port {port}...")
            driver = uc.Chrome(
                options=options,
                driver_executable_path=None, # Auto-download
                browser_executable_path=CHROME_PATH,
                version_main=None,
                port=port,
                suppress_welcome=True
            )
            if driver:
                break
        except Exception as e:
            print(f"[W{worker_id}] ❌ Start attempt {attempt+1} failed: {e}")
            if attempt < max_start_retries - 1:
                # Close potential orphan chrome from failed attempt
                try:
                    subprocess.run(f'taskkill /F /IM chrome.exe /FI "WINDOWTITLE eq tmp_{unique_id}*"', shell=True, capture_output=True)
                except: pass
                time.sleep(random.uniform(3, 7))
            else:
                raise e

    try:
        timezone_offset = -180  # Default to Moscow
        try:
            current_ip = check_current_ip(driver, worker_id)
            if current_ip and current_ip != 'Error':
                tz_info = get_timezone_for_ip(current_ip)
                timezone_offset = tz_info['offset']
                print(f"[W{worker_id}] 🌍 IP: {current_ip} → {tz_info['city']} (UTC{timezone_offset//60:+d})")
            else:
                print(f"[W{worker_id}] ⚠️ IP not detected, using Moscow TZ")
        except Exception as e:
            print(f"[W{worker_id}] ⚠️ Timezone detect failed: {e}")
        
        # Obsidian Stealth: Synchronize High-Entropy Client Hints
        full_version_list = ua_info['brand'].replace(';v=', ',').split(', ')
        brands_js = []
        for b in full_version_list:
            parts = b.split(';')
            if len(parts) == 2:
                name = parts[0].strip('"')
                ver = parts[1].replace('v=', '').strip('"')
                brands_js.append({'brand': name, 'version': ver})

        injection_script = f'''
            // Remove detectable overrides, let UC handle basic properties
            // Only spoof required environmental consistency
            Object.defineProperty(navigator, 'languages', {{get: () => ['ru-RU', 'ru', 'en-US', 'en']}});
            Date.prototype.getTimezoneOffset = function() {{ return {timezone_offset}; }};
            
            if (navigator.userAgentData) {{
                const original = navigator.userAgentData;
                Object.defineProperty(navigator, 'userAgentData', {{
                    get: () => ({{
                        ...original,
                        brands: {json.dumps(brands_js)},
                        mobile: false,
                        platform: "Windows",
                        getHighEntropyValues: (hints) => Promise.resolve({{
                            brands: {json.dumps(brands_js)},
                            mobile: false,
                            platform: "Windows",
                            platformVersion: "{ua_info['platform_version']}",
                            architecture: "{ua_info['architecture']}",
                            model: "{ua_info['model']}",
                            uaFullVersion: "{ua_info['full_ver']}",
                            bitness: "{ua_info['bitness']}",
                            fullVersionList: {json.dumps(brands_js)}
                        }})
                    }})
                }});
            }}
        '''
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': injection_script})
        
        # Pure UC: Let the browser handle headers naturally via User-Agent
        # No Network.setExtraHTTPHeaders
        
        # Simplified Fast Warming for Residential Proxies
        print(f"[W{worker_id}] 🧊 Fast Start...")
        try:
            print(f"[W{worker_id}] 🏠 Landing on Ozon Home...")
            driver.get("https://www.ozon.ru")
            time.sleep(random.uniform(5, 10))
            
            if "доступ ограничен" in driver.title.lower():
                 print(f"[W{worker_id}] ⚠️ Direct block on Ozon Home. Attempting search bypass...")
                 driver.get("https://www.ozon.ru/search/?text=iphone")
                 time.sleep(10)
            else:
                 print(f"[W{worker_id}] ✅ Ready!")
                 
        except Exception as e:
            print(f"[W{worker_id}] ⚠️ Warmup error: {e}")
            
        return driver, profile
    except Exception as e:
        print(f"[W{worker_id}] ❌ Failed to start UC: {e}")
        return None, None

def check_current_ip(driver, worker_id):
    """Robust IP detection with multi-service fallbacks."""
    services = [
        'https://api.ipify.org?format=json',
        'https://ipinfo.io/json',
        'https://httpbin.org/ip'
    ]
    
    for service in services:
        try:
            print(f"[W{worker_id}] 🔍 Checking IP via {service}...")
            driver.execute_script(f"window.open('{service}', '_blank');")
            time.sleep(2)
            driver.switch_to.window(driver.window_handles[-1])
            
            try:
                # Wait for any text to appear in body
                WebDriverWait(driver, 10).until(lambda d: d.find_element(By.TAG_NAME, 'body').text.strip() != '')
                body_text = driver.find_element(By.TAG_NAME, 'body').text
                
                # Cleanup before return
                driver.close()
                driver.switch_to.window(driver.window_handles[0])

                # Extract IP from common patterns
                ip_match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', body_text)
                if ip_match:
                    ip = ip_match.group(1)
                    return ip
            except Exception as e:
                print(f"[W{worker_id}] ⚠️ Service {service} failed or timed out: {e}")
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                continue
        except Exception as e:
            print(f"[W{worker_id}] ⚠️ IP check technical error ({service}): {e}")
            if len(driver.window_handles) > 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])

    return 'Error'

# Function attach_selenium_with_proxy removed as we use start_browser_uc directly

def extract_prices_api(session, sku, worker_id, cookies, ua):
    """
    Экстракция цен через API Ozon (быстрая версия).
    """
    try:
        product_link = f"/product/{sku}/"
        payload = {"url": product_link}
        headers = {
            "authority": "www.ozon.ru",
            "accept": "application/json",
            "accept-language": "ru-RU,ru;q=0.9",
            "user-agent": ua,
            "x-o3-app-name": "entrypoint-api",
            "x-o3-app-version": "master",
            "referer": f"https://www.ozon.ru{product_link}",
            "sec-ch-ua": '"Not_A Brand";v="124", "Chromium";v="124", "Google Chrome";v="124"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
        }

        response = session.get(
            API_ENDPOINT,
            params=payload,
            headers=headers,
            cookies=cookies,
            timeout=15
        )

        if response.status_code == 403:
            return {'status': 'ANTIBOT', 'is_antibot': True}
        if response.status_code != 200:
            return {'status': f'ERROR_{response.status_code}', 'is_antibot': False}

        data = response.json()
        widget_states = data.get("widgetStates", {})
        
        # Глубокий поиск в Nuxt/Composer структуре
        if not widget_states:
            try:
                vi = data.get("verticalInfo", {})
                composer = vi.get("composer", {}) or vi.get("pdp", {})
                widget_states = composer.get("widgetStates", {})
            except: pass

        price_card = None
        price_nocard = None
        price_old = None
        stock_status = "OK"
        product_name = "Unknown"
        
        # SEO для названия
        seo = data.get("seo") or data.get("SEO")
        if seo:
            product_name = seo.get("title") or "Unknown"

        price_widget_key = next((k for k in widget_states.keys() if "webPrice" in k), None)
        oos_widget_key = next((k for k in widget_states.keys() if "webOutOfStock" in k), None)
        
        if price_widget_key:
            try:
                price_state = json.loads(widget_states[price_widget_key])
                price_card = clean_price(price_state.get("cardPrice"))
                price_nocard = clean_price(price_state.get("price"))
                price_old = clean_price(price_state.get("originalPrice"))
                if "закончился" in str(price_state).lower():
                    stock_status = "OUT_OF_STOCK"
            except: pass

        if oos_widget_key:
             stock_status = "OUT_OF_STOCK"
             try:
                 oos_state = json.loads(widget_states[oos_widget_key])
                 if not price_nocard:
                     price_nocard = clean_price(oos_state.get("price"))
             except: pass
        
        # Fallback на текст
        if stock_status == "OK" and price_nocard is None:
            if "закончился" in str(data).lower():
                stock_status = "OUT_OF_STOCK"

        return {
            'price_card': price_card,
            'price_nocard': price_nocard,
            'price_old': price_old,
            'status': stock_status,
            'product_name': product_name,
            'is_antibot': False
        }

    except Exception as e:
        print(f"[W{worker_id}] ERROR: {e}")
        return {'status': 'ERROR', 'is_antibot': False}

def save_batch_to_db(batch):
    if not batch:
        return 0
    conn=psycopg2.connect(DB_URL)
    cur=conn.cursor()
    saved=0
    for item in batch:
        try:
            # Use SAVEPOINT to prevent transaction abortion on error
            cur.execute("SAVEPOINT sp1")
            # Clean SKU (remove .0 and ensure string)
            raw_sku = str(item['sku']).strip()
            if raw_sku.endswith('.0'):
                raw_sku = raw_sku[:-2]
                
            cur.execute("""
                INSERT INTO public.prices (sku, competitor_name, price_card, price_nocard, price_old, name, status, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (sku, competitor_name) 
                DO UPDATE SET 
                    price_card = EXCLUDED.price_card,
                    price_nocard = EXCLUDED.price_nocard,
                    price_old = EXCLUDED.price_old,
                    name = EXCLUDED.name,
                    status = EXCLUDED.status,
                    created_at = NOW()
            """, (
                raw_sku, 
                item['competitor_name'],
                item.get('price_card'), 
                item.get('price_nocard'), 
                item.get('price_old'), 
                item.get('product_name'), 
                item.get('status')
            ))
            cur.execute("RELEASE SAVEPOINT sp1")
            saved+=1
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp1")
            print(f"[DB ERROR] {item['sku']}: {e}")
    conn.commit()
    cur.close()
    conn.close()
    return saved

def run_single_batch(batch_products):
    """
    Обработка батча ПОСЛЕДОВАТЕЛЬНО (как в тесте).
    Один прогрев -> Один сеанс -> Прямое соединение без прокси.
    """
    global processed_count, results, global_cookies, global_ua
    processed_count = 0
    results = []
    start_time = time.time()
    
    # 1. Прогрев сессии (один раз на батч)
    global_cookies, global_ua = warmup_session()
    
    if not global_cookies or not global_ua:
        print("[ERROR] ❌ Не удалось прогреть сессию. Пропускаем батч.")
        return False

    # 2. Инициализация CFFI сессии с полученными куками (ПРЯМОЕ СОЕДИНЕНИЕ)
    session = CffiSession(impersonate="chrome124")
    # session.proxies = ... (УДАЛЕНО для прямого соединения)
    
    print(f"[OK] Сессия прогрета. Начинаем ПОСЛЕДОВАТЕЛЬНУЮ обработку {len(batch_products)} товаров...")

    for idx, (sku, name, competitor_name, sp_code) in enumerate(batch_products, 1):
        try:
            # Прямой вызов API
            res = extract_prices_api(session, sku, 0, global_cookies, global_ua)
            
            status = res.get('status', 'ERROR')
            
            # Сохранение результата
            result = {
                'sku': sku,
                'competitor_name': competitor_name or res.get('seller_name', 'Ozon'),
                'price_card': res.get('price_card'),
                'price_nocard': res.get('price_nocard'),
                'price_old': res.get('price_old'),
                'product_name': res.get('product_name'),
                'status': status
            }
            results.append(result)
            processed_count += 1

            if status == 'OK':
                print(f"[{idx}] ✅ SKU {sku}: {res.get('price_nocard')} руб.")
            elif status == 'OUT_OF_STOCK':
                print(f"[{idx}] 📦 SKU {sku}: Товар закончился")
            elif status == 'ANTIBOT':
                print(f"[{idx}] 🤖 ANTIBOT для SKU {sku}")
            else:
                print(f"[{idx}] ⚠️ ERROR для SKU {sku}: {status}")

            if processed_count % 20 == 0:
                print(f"\n📊 ПРОГРЕСС: {processed_count}/{len(batch_products)} товаров обработано\n")

            # Задержка между запросами (КАК В ТЕСТЕ - 0.5 сек)
            time.sleep(0.5)

        except Exception as e:
            print(f"[ERROR] SKU {sku}: {e}")
            time.sleep(1)
            
    elapsed = time.time() - start_time
    total = len(results)
    ok_count = sum(1 for r in results if r.get('status') == 'OK')
    out_of_stock = sum(1 for r in results if r.get('status') == 'OUT_OF_STOCK')
    antibot = sum(1 for r in results if r.get('status') == 'ANTIBOT')
    errors = sum(1 for r in results if r.get('status', '').startswith('ERROR'))
    
    if total > 0:
        print(f"\n{'='*100}")
        print(f"БАТЧ ЗАВЕРШЁН: {ok_count}/{total} товаров ({int(elapsed//60)}m {int(elapsed%60)}s)")
        print(f"  ✅ OK:                {ok_count:4d}")
        print(f"  📦 OUT_OF_STOCK:       {out_of_stock:4d}")
        print(f"  🤖 ANTIBOT:            {antibot:4d}")
        print(f"  ⚠️ ERRORS:             {errors:4d}")
        print(f"📊 СРЕДНЯЯ СКОРОСТЬ: {total/(elapsed/60):.1f} тов/мин")
        print(f"{'='*100}\n")
    return True
        
    elapsed = time.time() - start_time
    total = len(results)
    ok_count = sum(1 for r in results if r.get('status') == 'OK')
    out_of_stock = sum(1 for r in results if r.get('status') == 'OUT_OF_STOCK')
    antibot = sum(1 for r in results if r.get('status') == 'ANTIBOT')
    errors = sum(1 for r in results if r.get('status', '').startswith('ERROR'))
    
    if total > 0:
        print(f"\n{'='*100}")
        print(f"БАТЧ ЗАВЕРШЁН: {ok_count}/{total} товаров ({int(elapsed//60)}m {int(elapsed%60)}s)")
        print(f"  ✅ OK:                {ok_count:4d}")
        print(f"  📦 OUT_OF_STOCK:       {out_of_stock:4d}")
        print(f"  🤖 ANTIBOT:            {antibot:4d}")
        print(f"  ⚠️ ERRORS:             {errors:4d}")
        print(f"📊 СРЕДНЯЯ СКОРОСТЬ: {total/(elapsed/60):.1f} тов/мин")
        print(f"{'='*100}\n")
    return True

def load_proxies():
    # Now reading from upstreams.txt to single source truth
    try:
        with open('upstreams.txt','r') as f:
            line=f.readline().strip()
        parts=line.split(':')
        if len(parts)==4:
            proxy={'host':parts[0],'port':parts[1],'user':parts[2],'pass':parts[3]}
            print(f"[OK] MangoProxy template loaded: {parts[0]}:{parts[1]}")
            return proxy
    except Exception as e:
        print(f"[ERROR] Не удалось загрузить прокси: {e}")
    return None

def load_products_from_db():
    conn=psycopg2.connect(DB_URL)
    cur=conn.cursor()
    # Fetch competitor_name and sp_code as well
    cur.execute("""SELECT sku, name, competitor_name, sp_code FROM public.prices WHERE sku IS NOT NULL ORDER BY sku""")
    raw_data = cur.fetchall()
    
    products = []
    for (sku, name, comp_name, sp_code) in raw_data:
        # Clean SKU from float-like strings (e.g., '123.0' -> '123')
        clean_sku = str(sku).strip()
        if clean_sku.endswith('.0'):
            clean_sku = clean_sku[:-2]
        
        products.append((clean_sku, name or '', comp_name or '', sp_code or ''))
        
    cur.close()
    conn.close()
    return products

def generate_excel_report():
    print("\n[EXCEL] Generating report...")
    try:
        conn=psycopg2.connect(DB_URL)
        # Select ALL items, regardless of price
        query="""SELECT sku,name,competitor_name,price_card,price_nocard,price_old,status,sp_code FROM public.prices ORDER BY name, competitor_name"""
        df=pd.read_sql(query,conn)
        conn.close()
        if len(df)==0:
            print("[EXCEL] No data to report")
            return None
        
        # Data Cleaning
        df['competitor_name'] = df['competitor_name'].astype(str).str.strip()
        df['sp_code'] = df['sp_code'].astype(str).str.strip()
        print(f"\n[DEBUG] Raw Competitors from DB: {df['competitor_name'].unique()}")

        # Optional Mapping - but keep original if not found
        store_mapping={
            'Ссылка на наш магазин':'Наш магазин',
            'Магазин DeLonghi Group':'DeLonghi Group',
            'DeLonghi Group':'DeLonghi Group',
            'Delonghi Official Store':'DeLonghi Official',
            'Delonghi official store':'DeLonghi Official', # Case variant
            'DeLonghi Official Store':'DeLonghi Official'  # Case variant
        }
        # Use get(x, x) to keep original name if not in map
        df['competitor_name']=df['competitor_name'].map(lambda x:store_mapping.get(x,x))
        
        print(f"[DEBUG] Mapped Competitors: {df['competitor_name'].unique()}")

        # Apply status logic
        def fill_status(row):
            def check_val(val):
                if pd.isna(val): return True
                if str(val).lower().strip() in ['','none','nan']: return True
                return False

            status = str(row.get('status', '')).upper()
            p_card = row.get('price_card')
            p_nocard = row.get('price_nocard')
            
            # Text to display
            out_text = 'Товар закончился'
            
            # Condition 1: Explicit Status OOS
            if 'OUT_OF_STOCK' in status:
                # Keep last prices if they exist (now they are ints)
                return pd.Series([p_card, p_nocard, row.get('price_old'), out_text], 
                                index=['price_card', 'price_nocard', 'price_old', 'status'])
            
            # Condition 2: Missing Price but NOT Error
            if check_val(p_nocard):
                if 'BLOCKED' in status or 'ANTIBOT' in status:
                    text = 'Ошибка (Антибот)'
                    return pd.Series([text, text, text, text], index=['price_card', 'price_nocard', 'price_old', 'status'])
                elif 'ERROR' in status:
                    text = 'Ошибка парсинга'
                    return pd.Series([text, text, text, text], index=['price_card', 'price_nocard', 'price_old', 'status'])
                elif 'NO_PRICE' in status:
                    text = 'Нет цены'
                    return pd.Series([text, text, text, text], index=['price_card', 'price_nocard', 'price_old', 'status'])
                else:
                    return pd.Series([None, None, None, status], index=['price_card', 'price_nocard', 'price_old', 'status'])
                
            return pd.Series([p_card, p_nocard, row.get('price_old'), 'В наличии'], 
                            index=['price_card', 'price_nocard', 'price_old', 'status'])

        # Apply transformation
        df[['price_card', 'price_nocard', 'price_old', 'status']] = df.apply(fill_status, axis=1)

        # PRE-PROCESSING: Fill missing names per SP-CODE
        # For each SP-CODE, use the first non-empty name found across all stores
        def get_sp_name(sp_code):
            sp_data = df[df['sp_code'] == sp_code]['name']
            valid_names = sp_data.dropna()
            valid_names = valid_names[valid_names.astype(str).str.strip() != '']
            valid_names = valid_names[valid_names.astype(str).str.lower() != 'none']
            return valid_names.iloc[0] if len(valid_names) > 0 else None
        
        sp_name_map = {sp: get_sp_name(sp) for sp in df['sp_code'].unique() if sp}
        df['name'] = df.apply(lambda row: sp_name_map.get(row['sp_code']) if pd.isna(row['name']) or str(row['name']).strip() == '' else row['name'], axis=1)

        # Pivot the table
        # INDEX: ONLY SP_CODE (not name!) - one row per product
        # COLUMNS: COMPETITOR (Columns)
        # VALUES: SKU + NAME + PRICES
        # CRITICAL: Use dropna=False to preserve ALL stores even if they have sparse data
        pivot_df = df.pivot_table(
            index='sp_code',  # CHANGED: Only sp_code, not ['sp_code', 'name']
            columns='competitor_name', 
            values=['name', 'sku', 'price_card', 'price_nocard', 'price_old'],  # Added 'name'
            aggfunc='first',
            dropna=False  # THIS PRESERVES ALL COLUMNS!
        )
        
        # Swap levels to get Seller -> Attribute
        pivot_df.columns = pivot_df.columns.swaplevel(0, 1)
        
        # Rename attributes to Russian
        rename_map = {
            'name': 'Название',  # NEW: Add name column
            'sku': 'SKU',
            'price_card': 'Цена с картой',
            'price_nocard': 'Цена без карты',
            'price_old': 'Старая цена'
        }
        pivot_df = pivot_df.rename(columns=rename_map, level=1)
        
        # Leave empty cells blank (no "Нет данных" text)

        # Sort columns to group by Seller, then by Attribute order
        sellers = sorted(pivot_df.columns.get_level_values(0).unique())
        desired_order = ['Название', 'SKU', 'Цена с картой', 'Цена без карты', 'Старая цена']  # Added Название
        
        # Reindex columns
        new_columns = []
        for seller in sellers:
            for attr in desired_order:
                if (seller, attr) in pivot_df.columns:
                    new_columns.append((seller, attr))
        
        pivot_df = pivot_df.reindex(columns=new_columns)

        timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
        filename=f"ozon_prices_report_{timestamp}.xlsx"
        
        print(f"[EXCEL] Saving to {filename}...")
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Цены')
            worksheet = writer.sheets['Цены']
            
            # Formatting
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Green header style
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10, name='Roboto')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply styles to headers (rows 1 and 2)
            for row in worksheet.iter_rows(min_row=1, max_row=2):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
            
            # Auto-width
            for i, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze panes (skip 2 header rows)
            worksheet.freeze_panes = 'C3' 

        print(f"[EXCEL] Report created: {filename}")
        return filename
    except Exception as e:
        print(f"[EXCEL] Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def send_to_telegram(filename,stats_text):
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("[TG] Token or Chat ID not configured")
        return
    if not filename or not os.path.exists(filename):
        print("[TG] No file to send")
        return
    print("[TG] Sending to Telegram...")
    try:
        url=f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendDocument"
        with open(filename,'rb') as f:
            files={'document':f}
            data={'chat_id':TG_CHAT_ID}
            response=requests.post(url,data=data,files=files,timeout=60)
            print(f"[TG] Response status: {response.status_code}")  # DEBUG
            if response.status_code==200:
                print("[TG] Report sent successfully")
                # Даем системе время закрыть дескриптор файла
                time.sleep(5)
                try:
                    if os.path.exists(filename):
                        os.remove(filename)
                        print(f"[TG] ✅ File deleted: {filename}")
                except Exception as del_err:
                    print(f"[TG] ❌ Failed to delete file: {del_err}")
            else:
                print(f"[TG] Error: {response.text}")
    except Exception as e:
        print(f"[TG] Error sending file: {e}")

def kill_all_browsers():
    import subprocess
    import glob
    try:
        # Kill all Chrome and Chromedriver instances
        print("[CLEANUP] 🔪 Killing Chrome and Chromedriver processes...")
        subprocess.run('taskkill /F /IM chrome.exe /T', shell=True, capture_output=True, timeout=10)
        subprocess.run('taskkill /F /IM chromedriver.exe /T', shell=True, capture_output=True, timeout=10)
        
        # Kill all orphan python workers
        current_pid = os.getpid()
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                if proc.info['name'] == 'python.exe' and proc.info['pid'] != current_pid:
                    # Only kill if it looks like a worker or part of this project
                    # For safety, we kill all other python processes in this context
                    psutil.Process(proc.info['pid']).kill()
            except:
                pass
        
        # Clear Logs
        print("[CLEANUP] 📝 Clearing logs and temporary files...")
        if os.path.exists("proxy_log.txt"):
            with open("proxy_log.txt", "w") as f:
                f.write(f"--- Log cleared at {datetime.now()} ---\n")
        
        # Clear Debug HTML files
        debug_files = glob.glob("debug_html_*.html")
        for f in debug_files:
            try:
                os.remove(f)
            except:
                pass
                
        print("[CLEANUP] ✅ Все процессы убиты, логи очищены.")
    except Exception as e:
        print(f"[CLEANUP] ⚠️ Ошибка при выполнении очистки: {e}")

def main():
    global processed_count,results,last_processed_skus,batch_complete
    print("="*100)
    print("OZON PRODUCTION PARSER - РЕЖИМ НЕПРЕРЫВНЫХ БАТЧЕЙ")
    print(f"СТРАТЕГИЯ: {MAX_PRODUCTS_PER_BATCH} товаров -> сохранить -> БЕЗ ПАУЗ -> следующие {MAX_PRODUCTS_PER_BATCH}")
    print(f"СКОРОСТЬ: Максимальная! Без задержек между батчами!")
    print("="*100)
    batch_number=1
    total_parsed=0
    all_products=load_products_from_db()
    print(f"\n[INIT] Загружено {len(all_products)} товаров")
    print("[RANDOMIZE] Перемешиваем товары для естественного поведения...")
    random.shuffle(all_products)
    print("[RANDOMIZE] OK - Товары перемешаны!\n")
    current_offset=0
    while current_offset<len(all_products):
        print(f"\n{'='*100}")
        print(f"\n[BATCH #{batch_number}] Товары {current_offset+1} - {min(current_offset+MAX_PRODUCTS_PER_BATCH,len(all_products))}")
        print(f"{'='*100}\n")
        
        batch_products=all_products[current_offset:current_offset+MAX_PRODUCTS_PER_BATCH]
        print(f"[INIT] Загружено {len(batch_products)} товаров для обработки\n")
        
        success=run_single_batch(batch_products)
        if not success:
            print("[ERROR] Ошибка при обработке батча")
            break
            
        batch_processed=len(results)
        total_parsed+=batch_processed
        
        print(f"\n{'='*100}")
        print(f"✅ БАТЧ #{batch_number} ЗАВЕРШЁН")
        print(f"   Обработано: {batch_processed} товаров")
        print(f"   Всего обработано: {total_parsed}/{len(all_products)}")
        print(f"{'='*100}\n")
        
        print(f"[CLEANUP] 🔪 Очистка после батча...")
        kill_all_browsers()
        
        if results:
            print(f"\n[DB] 💾 Сохранение {len(results)} товаров в базу данных...")
            saved=save_batch_to_db(results)
            print(f"[DB] ✅ Сохранено {saved} товаров")
            
        current_offset+=MAX_PRODUCTS_PER_BATCH
        if current_offset>=len(all_products):
            print("[COMPLETE] Все товары обработаны!")
            break
            
        print(f"{'='*100}")
        print(f"🚀 СЛЕДУЮЩИЙ БАТЧ: {current_offset+1} - {min(current_offset+MAX_PRODUCTS_PER_BATCH,len(all_products))}")
        print(f"   ПАУЗА ОХЛАЖДЕНИЯ СЕССИИ: 5 секунд...")
        print(f"{'='*100}\n")
        time.sleep(5)
        batch_number+=1
        # DISABLE profile deletion for persistence during troubleshooting
        # print(f"[CLEANUP] 🗑️ Удаляем ВСЕ профили Chrome...")
        # clean_old_chrome_profiles(max_age_minutes=0)
        # time.sleep(2)
        # profiles_dir=Path("C:/Temp/chrome_profiles/ozon")
        # remaining_profiles=list(profiles_dir.glob("p*")) if profiles_dir.exists() else []
        # if remaining_profiles:
        #     print(f"[WARNING] ⚠️ Осталось {len(remaining_profiles)} профилей! Удаляем повторно...")
        #     for profile in remaining_profiles:
        #         try:
        #             shutil.rmtree(profile,ignore_errors=True)
        #         except:
        #             pass
        #     time.sleep(2)
        # else:
        #     print(f"[CLEANUP] ✅ Все профили удалены")
        if results:
            print(f"\n[DB] 💾 Сохранение {len(results)} товаров в базу данных...")
            saved=save_batch_to_db(results)
            print(f"[DB] ✅ Сохранено {saved} товаров")
        if current_offset+MAX_PRODUCTS_PER_BATCH>=len(all_products):
            print("[COMPLETE] Все товары обработаны!")
            break
        current_offset+=MAX_PRODUCTS_PER_BATCH
        print(f"{'='*100}")
        print(f"🚀 СЛЕДУЮЩИЙ БАТЧ: {current_offset+1} - {min(current_offset+MAX_PRODUCTS_PER_BATCH,len(all_products))}")
        print(f"   ПАУЗА ОХЛАЖДЕНИЯ СЕССИИ: 5 секунд...")
        print(f"{'='*100}\n")
        time.sleep(5)
        batch_number+=1

    print(f"\n{'='*100}")
    print(f"\n[COMPLETE] ВСЕ БАТЧИ ЗАВЕРШЕНЫ!")
    print(f"   Всего батчей: {batch_number}")
    print(f"   Всего товаров обработано: {total_parsed}/{len(all_products)}")
    print(f"{'='*100}\n")
    print("[INFO] Проверяем сколько цен в базе данных...")
    try:
        conn=psycopg2.connect(DB_URL)
        cur=conn.cursor()
        cur.execute("SELECT COUNT(*) FROM public.prices WHERE price_card IS NOT NULL")
        count=cur.fetchone()[0]
        print(f"[INFO] ✅ Цен в БД: {count}/{len(all_products)}")
        conn.close()
    except Exception as e:
        print(f"[ERROR] Не удалось проверить БД: {e}")
    print("\n"+"="*100)
    print("📊 ГЕНЕРАЦИЯ ОТЧЕТА И ОТПРАВКА В TELEGRAM")
    print("="*100)
    try:
        excel_file=generate_excel_report()
        if excel_file:
            print(f"[REPORT] ✅ Отчет создан: {excel_file}")
            print("[TELEGRAM] 📤 Отправка отчета в Telegram...")
            send_to_telegram(excel_file,"")
            print("[TELEGRAM] ✅ Отчет отправлен!")
        else:
            print("[REPORT] ⚠️ Не удалось создать отчет")
            
        # VIOLATION CHECK
        print("\n"+"="*100)
        print("🕵️ ПРОВЕРКА НАРУШЕНИЙ (Скрины + Telegram)")
        print("="*100)
        if check_violations is not None:
            check_violations.run_check()
        else:
            print("[INFO] Модуль нарушений (check_violations) не найден, пропускаем...")
            
    except Exception as e:
        print(f"[ERROR] Ошибка при генерации/отправке отчета: {e}")
        import traceback
        traceback.print_exc()
    print("\n"+"="*100)
    print("✅ ВСЁ ГОТОВО!")
    print("="*100+"\n")

if __name__=='__main__':
    print("DEBUG: Starting parser...")
    print("\n"+"="*70)
    print("ВАЖНО: Убедитесь что 3proxy запущен!")
    print("="*70)
    print("\nЕсли 3proxy НЕ запущен:")
    print("   1. Откройте новое окно терминала")
    print("   2. Запустите: start_3proxy.bat")
    print("   3. Дождитесь сообщения '3proxy started'")
    print("\n3proxy должен слушать на 127.0.0.1:8118 (SOCKS5)")
    print("="*70)
    time.sleep(2)
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n[STOP] Ostanovleno polzovatelem")
    except Exception as e:
        print(f"\n\n[ERROR]: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n" + "="*70)
        print("ПАРСИНГ ЗАВЕРШЁН - Выполнение финальной очистки...")
        print("="*70)
        
        # Comprehensive cleanup: Kill workers, close browsers, clear logs
        kill_all_browsers()
        
        # Terminate proxy server (3proxy / auth_forwarder)
        try:
            import subprocess
            print("[CLEANUP] Останавливаем прокси-сервер (3proxy/Forwarder)...")
            subprocess.run('taskkill /F /IM 3proxy.exe /T', shell=True, capture_output=True, timeout=5)
            # Kill the proxy CMD window by its title set in start_3proxy.bat
            subprocess.run('taskkill /F /FI "WINDOWTITLE eq OzonProxyForwarder*" /T', shell=True, capture_output=True, timeout=5)
            print("[CLEANUP] ✅ Прокси-сервер остановлен")
        except Exception as e:
            print(f"[CLEANUP] ⚠️ Не удалось остановить прокси: {e}")
        
        print("\n" + "="*70)
        print("✅ ВСЕ ЦЕПОЧКИ ЗАКРЫТЫ. ВЫХОД.")
        print("="*70)
        sys.exit(0)
