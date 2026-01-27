
import os
import psycopg2
import pandas as pd
import requests
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

DB_URL = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASS')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
TG_BOT_TOKEN = os.getenv('TG_BOT_TOKEN')
TG_CHAT_ID = os.getenv('TG_CHAT_ID')

def generate_wb_report():
    print("\n[EXCEL] Generating WB report...")
    try:
        conn = psycopg2.connect(DB_URL)
        query = """
            SELECT sku, competitor_name, price_card, price_nocard, price_old, status 
            FROM public.wb_prices 
            ORDER BY sku, competitor_name
        """
        df = pd.read_sql(query, conn)
        conn.close()
        
        if len(df) == 0:
            print("[EXCEL] No WB data to report")
            return None
        
        # Create report rows - one row per SKU per seller
        report_rows = []
        
        for _, row in df.iterrows():
            sku = row['sku']
            seller = row['competitor_name']
            price_card = row['price_card']
            price_nocard = row['price_nocard']
            price_old = row['price_old']
            status = row.get('status')
            
            # Format values
            def format_price(val, st):
                if st == 'OUT_OF_STOCK': 
                    return 'Товар закончился'
                if pd.notna(val) and val != 0: 
                    return val
                if st == 'ANTIBOT': 
                    return 'Ошибка (Антибот)'
                if st == 'ERROR': 
                    return 'Ошибка парсинга'
                return ''
            
            # Create row
            report_row = {
                'SKU': sku,
                'Продавец': seller,
                'Промо': format_price(price_card, status),
                'Цена': format_price(price_nocard, status),
                'Старая': format_price(price_old, status)
            }
            
            report_rows.append(report_row)
        
        result_df = pd.DataFrame(report_rows)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"wb_prices_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Цены WB', index=False)
            ws = writer.sheets['Цены WB']
            
            # Styles
            header_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid") # Purple for WB
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply header styles
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Set column widths
            ws.column_dimensions['A'].width = 15  # SKU
            ws.column_dimensions['B'].width = 35  # Продавец
            ws.column_dimensions['C'].width = 15  # Промо
            ws.column_dimensions['D'].width = 15  # Цена
            ws.column_dimensions['E'].width = 15  # Старая
            
            # Apply borders to all cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Freeze panes
            ws.freeze_panes = 'A2'
            
            # Add filters
            ws.auto_filter.ref = ws.dimensions
                
        print(f"[EXCEL] ✅ WB Report created: {filename}")
        return filename
    except Exception as e:
        print(f"[EXCEL] ❌ Error: {e}")
        return None

def send_wb_report(filename):
    if not filename: return
    print(f"[TELEGRAM] 📤 Sending {filename}...")
    
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("[TELEGRAM] ❌ Token or Chat ID missing")
        return
        
    url = f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendDocument"
    try:
        with open(filename, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': TG_CHAT_ID, 'caption': f'🟣 Отчёт по ценам Wildberries\n\n✅ Файл: {filename}'}
            resp = requests.post(url, files=files, data=data, timeout=30)
        
        # Проверка результата и удаление файла (после закрытия файла)
        if resp.status_code == 200 and resp.json().get('ok'):
            print("[TG] ✅ Report sent successfully")
            
            # Удаляем файл после успешной отправки
            try:
                os.remove(filename)
                print(f"[TG] 🗑️  File {filename} deleted")
            except Exception as e:
                print(f"[TG] ⚠️  Could not delete file: {e}")
        else:
            print(f"[TG] ❌ Failed: {resp.text}")
    except Exception as e:
        print(f"[TG] ❌ Error: {e}")

if __name__ == "__main__":
    f = generate_wb_report()
    send_wb_report(f)
