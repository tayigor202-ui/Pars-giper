import os
import psycopg2
import pandas as pd
import requests
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

DB_URL = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASS')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
TG_BOT_TOKEN = os.getenv('TG_BOT_TOKEN')
TG_CHAT_ID = os.getenv('TG_CHAT_ID')

def generate_lemana_report():
    print("\n[EXCEL] Generating Lemana report...")
    try:
        conn = psycopg2.connect(DB_URL)
        query = """
            SELECT 
                sku as "SKU", 
                name as "Наименование", 
                competitor_name as "Конкурент", 
                price as "Цена", 
                ric_leroy_price as "РИЦ Леруа",
                CASE WHEN violation_detected THEN 'ДА' ELSE 'Нет' END as "Нарушение",
                violation_screenshot as "Скриншот",
                sp_code as "СП-Код", 
                last_updated as "Дата обновления"
            FROM public.lemana_prices 
            ORDER BY sp_code, competitor_name
        """
        df = pd.read_sql(query, conn)
        conn.close()
        
        if len(df) == 0:
            print("[EXCEL] No Lemana data to report")
            return None
        
        # Add a clickable link for screenshots if they exist
        # We assume local file access for now, or we could prefix with a base URL
        # For simplicity, we just keep the path
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"lemana_prices_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Цены Lemana', index=False)
            ws = writer.sheets['Цены Lemana']
            
            # Styles
            header_fill = PatternFill(start_color="008C45", end_color="008C45", fill_type="solid") # Green for Lemana (Leroy)
            header_font = Font(bold=True, color="FFFFFF", size=11)
            violation_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light red
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply header styles
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                
            # Set column widths
            ws.column_dimensions['A'].width = 15  # SKU
            ws.column_dimensions['B'].width = 45  # Name
            ws.column_dimensions['C'].width = 20  # Competitor
            ws.column_dimensions['D'].width = 12  # Price
            ws.column_dimensions['E'].width = 12  # RIC
            ws.column_dimensions['F'].width = 12  # Violation
            ws.column_dimensions['G'].width = 40  # Screenshot
            ws.column_dimensions['H'].width = 15  # SP-KOD
            ws.column_dimensions['I'].width = 20  # Date
            
            # Apply borders and highlight violations
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9), start=2):
                is_violation = ws.cell(row=row_idx, column=6).value == 'ДА'
                for cell in row:
                    cell.border = border
                    if is_violation:
                        cell.fill = violation_fill
                
                # Make screenshot cell a hyperlink if it has a value
                screenshot_cell = ws.cell(row=row_idx, column=7)
                if screenshot_cell.value:
                    val = screenshot_cell.value
                    # If it's a relative path, we can try to make it a link
                    # For now keep it as text but highlighted
                    screenshot_cell.font = Font(color="0000FF", underline="single")
            
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
                
        print(f"[EXCEL] ✅ Lemana Report created: {filename}")
        return filename
    except Exception as e:
        print(f"[EXCEL] ❌ Error: {e}")
        return None

def send_lemana_report(filename):
    if not filename: return
    print(f"[TELEGRAM] 📤 Sending {filename}...")
    
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("[TELEGRAM] ❌ Token or Chat ID missing")
        return
        
    url = f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendDocument"
    try:
        with open(filename, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': TG_CHAT_ID, 'caption': f'🟢 Отчёт по ценам Lemana Pro\n\n✅ Файл: {filename}'}
            resp = requests.post(url, files=files, data=data, timeout=30)
        
        if resp.status_code == 200 and resp.json().get('ok'):
            print("[TG] ✅ Report sent successfully")
            try:
                os.remove(filename)
            except: pass
        else:
            print(f"[TG] ❌ Failed: {resp.text}")
    except Exception as e:
        print(f"[TG] ❌ Error: {e}")

if __name__ == "__main__":
    f = generate_lemana_report()
    send_lemana_report(f)
