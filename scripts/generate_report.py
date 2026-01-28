
import os
import psycopg2
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Load environment variables
load_dotenv()

# Database configuration
DB_URL = os.getenv('DB_URL')
if not DB_URL:
    DB_USER = os.getenv('DB_USER')
    DB_PASS = os.getenv('DB_PASS')
    DB_HOST = os.getenv('DB_HOST')
    DB_PORT = os.getenv('DB_PORT')
    DB_NAME = os.getenv('DB_NAME')
    DB_URL = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

def generate_excel_report():
    print("\n[EXCEL] Generating report...")
    try:
        conn = psycopg2.connect(DB_URL)
        # Select ALL items with price details
        query = """
            SELECT sku, competitor_name, price_card, price_nocard, price_old, status 
            FROM public.prices 
            ORDER BY sku, competitor_name
        """
        df = pd.read_sql(query, conn)
        conn.close()
        
        if len(df) == 0:
            print("[EXCEL] No data to report")
            return None
        
        # Mapping technical names to readable store names
        store_mapping = {
            'Ссылка на наш магазин': 'Наш магазин',
            'DeLonghi Group': 'DeLonghi Group',
            'Delonghi Official Store': 'DeLonghi Official'
        }
        df['competitor_name'] = df['competitor_name'].map(lambda x: store_mapping.get(x, x))
        
        print("\n[DEBUG] Competitors in Report DF:")
        print(df['competitor_name'].unique())
        
        # Create report rows - one row per SKU per seller
        report_rows = []
        
        for _, row in df.iterrows():
            sku = row['sku']
            seller = row['competitor_name']
            price_card = row['price_card']
            price_nocard = row['price_nocard']
            price_old = row['price_old']
            status = row.get('status')
            
            # Format values
            def format_price(val, item_status):
                # PRIORITY 1: Explicit Status
                if item_status == 'OUT_OF_STOCK':
                    return 'Товар закончился'
                
                # PRIORITY 2: Explicit Text in Price
                if isinstance(val, str) and 'закончился' in val.lower():
                    return 'Товар закончился'
                    
                # PRIORITY 3: Value exists
                if pd.notna(val) and val != 0:
                    return val
                    
                # PRIORITY 4: Fallback Status
                if item_status == 'ANTIBOT':
                    return 'Ошибка (Антибот)'
                if item_status == 'ERROR':
                    return 'Ошибка парсинга'
                if item_status == 'NO_PRICE':
                    return 'Нет цены'
                return ''
            
            # Create row
            report_row = {
                'SKU': sku,
                'Продавец': seller,
                'Промо': format_price(price_card, status),
                'Цена': format_price(price_nocard, status),
                'Старая': format_price(price_old, status)
            }
            
            report_rows.append(report_row)
        
        result_df = pd.DataFrame(report_rows)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ozon_prices_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Цены', index=False)
            worksheet = writer.sheets['Цены']
            
            # Styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Apply header styles
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 15  # SKU
            worksheet.column_dimensions['B'].width = 35  # Продавец
            worksheet.column_dimensions['C'].width = 15  # Промо
            worksheet.column_dimensions['D'].width = 15  # Цена
            worksheet.column_dimensions['E'].width = 15  # Старая
            
            # Apply borders to all cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Freeze panes
            worksheet.freeze_panes = 'A2'
            
            # Add filters
            worksheet.auto_filter.ref = worksheet.dimensions
        
        print(f"[EXCEL] ✅ Report created: {filename}")
        print(f"[EXCEL] 📊 Rows: {len(result_df)}, Columns: {len(result_df.columns)}")
        return filename
    except Exception as e:
        print(f"[EXCEL] ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    print("=" * 60)
    print("ГЕНЕРАЦИЯ ОТЧЁТА OZON")
    print("=" * 60)
    filename = generate_excel_report()
    if filename:
        print(f"\n✅ Отчёт успешно создан: {filename}")
        os.system(f"start {filename}") # Open file automatically
    else:
        print("\n❌ Не удалось создать отчёт")
