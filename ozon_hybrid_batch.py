#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch Hybrid Ozon Parser - для обработки множества товаров

Оптимизирован для пакетной обработки:
- Один прогрев браузера на всю сессию
- Переиспользование curl-cffi сессии
- Сохранение результатов в JSON и Excel
"""
import json
import time
import os
import re
from datetime import datetime
from curl_cffi.requests import Session as CffiSession
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIGURATION ---
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
PROXY_SERVER = "127.0.0.1:8118"
API_ENDPOINT = "https://www.ozon.ru/api/composer-api.bx/page/json/v2"

def warmup_session_once():
    """
    Прогрев сессии один раз для всей пакетной обработки.
    Возвращает: (cookies_dict, user_agent_string)
    """
    options = uc.ChromeOptions()
    options.binary_location = CHROME_PATH
    options.add_argument(f"--proxy-server=http://{PROXY_SERVER}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=ru-RU")
    
    driver = None
    try:
        print("[WARMUP] Запуск браузера для прогрева сессии...")
        driver = uc.Chrome(options=options, browser_executable_path=CHROME_PATH)
        
        print("[WARMUP] Посещение главной страницы Ozon...")
        driver.get("https://www.ozon.ru")
        time.sleep(5)
        
        # Посещаем любую страницу товара для полного прогрева
        print("[WARMUP] Прогрев на странице товара...")
        driver.get("https://www.ozon.ru/product/1067025156/")
        time.sleep(8)
        
        # Извлекаем куки и User-Agent
        selenium_cookies = driver.get_cookies()
        cookies_dict = {cookie['name']: cookie['value'] for cookie in selenium_cookies}
        user_agent = driver.execute_script("return navigator.userAgent;")
        
        print(f"[WARMUP] ✅ Сессия прогрета. Извлечено {len(cookies_dict)} куки.")
        
        return cookies_dict, user_agent
        
    except Exception as e:
        print(f"[ERROR] Ошибка прогрева: {e}")
        return None, None
    finally:
        if driver:
            driver.quit()
            print("[WARMUP] Браузер закрыт.")

def fetch_product_batch(product_links, cookies_dict, user_agent):
    """
    Обработка пакета товаров через API с одной сессией curl-cffi.
    """
    session = CffiSession(impersonate="chrome124")
    
    headers = {
        "authority": "www.ozon.ru",
        "accept": "application/json",
        "accept-language": "ru-RU,ru;q=0.9",
        "user-agent": user_agent,
        "x-o3-app-name": "entrypoint-api",
        "x-o3-app-version": "master",
        "sec-ch-ua": '"Not_A Brand";v="124", "Chromium";v="124", "Google Chrome";v="124"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }
    
    proxies = {
        "http": f"http://{PROXY_SERVER}",
        "https": f"http://{PROXY_SERVER}",
    }
    
    results = []
    
    for idx, product_link in enumerate(product_links, 1):
        try:
            print(f"[{idx}/{len(product_links)}] Обработка: {product_link}")
            
            payload = {"url": product_link}
            headers["referer"] = f"https://www.ozon.ru{product_link}"
            
            response = session.get(
                API_ENDPOINT,
                params=payload,
                headers=headers,
                cookies=cookies_dict,
                proxies=proxies,
                timeout=30
            )
            
            if response.status_code != 200:
                print(f"  ❌ Ошибка: HTTP {response.status_code}")
                results.append({
                    "link": product_link,
                    "status": f"ERROR_{response.status_code}",
                    "title": None,
                    "price": None
                })
                continue
            
            data = response.json()
            seo_data = data.get("seo") or data.get("SEO")
            
            if seo_data:
                # Извлечение цены из JSON-LD если нужно
                price = seo_data.get("price")
                currency = seo_data.get("currency", "RUB")
                
                if price is None and "script" in seo_data:
                    for script in seo_data["script"]:
                        if "application/ld+json" in script.get("type", ""):
                            try:
                                ld_json = json.loads(script.get("innerHTML", "{}"))
                                if isinstance(ld_json, list):
                                    ld_json = ld_json[0]
                                offers = ld_json.get("offers", {})
                                if isinstance(offers, list):
                                    offers = offers[0]
                                price = offers.get("price")
                                currency = offers.get("priceCurrency", currency)
                                break
                            except:
                                pass
                
                title = seo_data.get("title") or "Unknown"
                print(f"  ✅ {title[:50]}... - {price} {currency}")
                
                results.append({
                    "link": product_link,
                    "status": "OK",
                    "title": title,
                    "price": price,
                    "currency": currency
                })
            else:
                print(f"  ⚠️ SEO ключ не найден")
                results.append({
                    "link": product_link,
                    "status": "NO_SEO",
                    "title": None,
                    "price": None
                })
            
            # Небольшая задержка между запросами
            time.sleep(1)
            
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            results.append({
                "link": product_link,
                "status": "EXCEPTION",
                "title": None,
                "price": None,
                "error": str(e)
            })
    
    return results

def save_results(results, filename="hybrid_results"):
    """Сохранение результатов в JSON и Excel."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # JSON
    json_file = f"{filename}_{timestamp}.json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\n💾 Сохранено в JSON: {json_file}")
    
    # Excel
    excel_file = f"{filename}_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Заголовки
    headers = ["№", "Ссылка", "Статус", "Название", "Цена", "Валюта"]
    ws.append(headers)
    
    # Стиль заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    for idx, result in enumerate(results, 1):
        ws.append([
            idx,
            result.get("link", ""),
            result.get("status", ""),
            result.get("title", ""),
            result.get("price", ""),
            result.get("currency", "")
        ])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_file)
    print(f"💾 Сохранено в Excel: {excel_file}")
    
    return json_file, excel_file

def process_batch(product_links):
    """Главная функция пакетной обработки."""
    print("="*70)
    print(f" HYBRID BATCH PARSER - {len(product_links)} товаров")
    print("="*70)
    
    start_time = time.time()
    
    # Шаг 1: Прогрев сессии
    cookies, user_agent = warmup_session_once()
    
    if not cookies or not user_agent:
        print("❌ Не удалось прогреть сессию!")
        return
    
    # Шаг 2: Обработка всех товаров
    results = fetch_product_batch(product_links, cookies, user_agent)
    
    # Шаг 3: Сохранение результатов
    json_file, excel_file = save_results(results)
    
    elapsed = time.time() - start_time
    
    # Статистика
    success_count = sum(1 for r in results if r.get("status") == "OK")
    error_count = len(results) - success_count
    
    print("\n" + "="*70)
    print(" ИТОГИ ")
    print("="*70)
    print(f"✅ Успешно:  {success_count}/{len(results)}")
    print(f"❌ Ошибки:   {error_count}/{len(results)}")
    print(f"⏱️  Время:    {elapsed:.1f}s ({elapsed/len(results):.1f}s на товар)")
    print(f"📊 Скорость:  {len(results)/(elapsed/60):.1f} товаров/мин")
    print("="*70)

if __name__ == "__main__":
    import sys
    
    # Вариант 1: Чтение из файла
    if os.path.exists("test_products_10.txt"):
        with open("test_products_10.txt", "r", encoding="utf-8") as f:
            product_links = [line.strip() for line in f if line.strip()]
        print(f"📂 Загружено {len(product_links)} товаров из test_products_10.txt\n")
    else:
        # Вариант 2: Тестовый список
        product_links = [
            "/product/1067025156/",
            "/product/1564586312/",
            "/product/1401683802/",
        ]
        print(f"⚠️  Используется тестовый список: {len(product_links)} товаров\n")
    
    if len(product_links) > 0:
        process_batch(product_links)
    else:
        print("❌ Список товаров пуст!")
